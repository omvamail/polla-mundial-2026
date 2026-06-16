import os
import json
import openpyxl

def get_match_date(group_letter, match_index):
    groups = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
    if group_letter not in groups:
        return '2026-06-11'
        
    g_idx = groups.index(group_letter)
    
    day_offset = 0
    if match_index in [0, 1]:
        day_offset = g_idx // 2
    elif match_index in [2, 3]:
        day_offset = 6 + (g_idx // 2)
    else:
        day_offset = 12 + (g_idx // 2)
        
    day = 11 + day_offset
    return f"2026-06-{day:02d}"

def main():
    wb = openpyxl.load_workbook('Marcadores Mundial 2026.xlsx', data_only=True)
    res_sheet = wb['Resultados']
    
    matches_list = []
    group_match_counts = {}
    date_match_counts = {}
    
    # Read matches
    for r in range(2, res_sheet.max_row + 1):
        g = res_sheet.cell(r, 1).value
        local = res_sheet.cell(r, 2).value
        visitante = res_sheet.cell(r, 6).value
        
        if g and local and visitante:
            g = str(g).strip()
            local = str(local).strip()
            visitante = str(visitante).strip()
            
            # Count match index in group (0 to 5)
            group_match_counts[g] = group_match_counts.get(g, 0) + 1
            m_idx = group_match_counts[g] - 1
            
            date_str = get_match_date(g, m_idx)
            
            # Count match index in that specific date (0 to 3)
            date_match_counts[date_str] = date_match_counts.get(date_str, 0) + 1
            day_idx = date_match_counts[date_str] - 1
            
            # Assign hours sequentially: 12:00, 15:00, 18:00, 21:00
            hours = ['12:00', '15:00', '18:00', '21:00']
            time_str = hours[day_idx] if day_idx < len(hours) else '22:00'
            
            matches_list.append({
                'row': r,
                'group': g,
                'local': local,
                'visitante': visitante,
                'date': date_str,
                'time': time_str,
                'predictions': {}
            })
            
    # Read predictions
    participants = ['Deivid', 'Cesar', 'Ferney', 'Edisabet', 'Jhair', 'Duvan']
    for p in participants:
        p_sheet = wb[p]
        for m in matches_list:
            r = m['row']
            pred = p_sheet.cell(r, 5).value
            if pred:
                pred = str(pred).strip()
            else:
                pred = ""
            m['predictions'][p] = pred

    os.makedirs('static', exist_ok=True)
    
    with open('static/data.json', 'w', encoding='utf-8') as f:
        json.dump(matches_list, f, indent=4, ensure_ascii=False)
        
    print(f"Exported {len(matches_list)} matches with predictions, dates, and times successfully to static/data.json!")

if __name__ == '__main__':
    main()
