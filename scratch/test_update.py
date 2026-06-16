import openpyxl
import shutil

# Copy original file to a temp file for testing
shutil.copyfile('Marcadores Mundial 2026.xlsx', 'Marcadores_Test.xlsx')

try:
    # 1. Load temp file with data_only=False (to preserve formulas)
    wb = openpyxl.load_workbook('Marcadores_Test.xlsx', data_only=False)
    sheet = wb['Resultados']
    
    # Let's inspect original value and formula
    row = 4 # Row 4: República Checa vs Sudáfrica (currently no scores)
    print("Before update:")
    print("Row 4, Col C (Goles Local):", sheet.cell(row, 3).value)
    print("Row 4, Col E (Goles Visitante):", sheet.cell(row, 5).value)
    print("Row 4, Col G (Resultado):", sheet.cell(row, 7).value)
    
    # 2. Update scores in Resultados
    sheet.cell(row, 3).value = 3 # 3 goals for local
    sheet.cell(row, 5).value = 1 # 1 goal for visitor
    sheet.cell(row, 7).value = "Local (L)" # Outcome
    
    # 3. Save
    wb.save('Marcadores_Test.xlsx')
    print("\nSaved updated Excel.")
    
    # 4. Load it back with data_only=False and check if formulas in Deivid sheet are intact
    wb2 = openpyxl.load_workbook('Marcadores_Test.xlsx', data_only=False)
    d_sheet = wb2['Deivid']
    print("\nAfter update (data_only=False):")
    print("Deivid sheet Row 4 Col F (Acierto Formula):", d_sheet.cell(row, 6).value)
    print("Deivid sheet Row 4 Col G (Puntos Formula):", d_sheet.cell(row, 7).value)
    print("Deivid sheet Cell J1 (Total Formula):", d_sheet.cell(1, 10).value)
    
    # 5. Load it back with data_only=True and see what's loaded (values will be cached unless opened in Excel, let's see what openpyxl shows)
    wb3 = openpyxl.load_workbook('Marcadores_Test.xlsx', data_only=True)
    d_sheet3 = wb3['Deivid']
    print("\nAfter update (data_only=True - cached values from last save):")
    print("Deivid sheet Row 4 Col F (Acierto Value):", d_sheet3.cell(row, 6).value)
    print("Deivid sheet Row 4 Col G (Puntos Value):", d_sheet3.cell(row, 7).value)
    print("Deivid sheet Cell J1 (Total Value):", d_sheet3.cell(1, 10).value)
    
except Exception as e:
    print("Error:", e)
