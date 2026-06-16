import openpyxl

wb = openpyxl.load_workbook('Marcadores Mundial 2026.xlsx', data_only=True)
sheet = wb['Resultados']
teams = set()
for r in range(2, sheet.max_row + 1):
    local = sheet.cell(r, 2).value
    visit = sheet.cell(r, 6).value
    if local:
        teams.add(local.strip())
    if visit:
        teams.add(visit.strip())

print("Unique Teams:")
for t in sorted(list(teams)):
    print(f"'{t}'")
