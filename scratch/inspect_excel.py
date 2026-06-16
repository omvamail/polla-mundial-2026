import openpyxl

def main():
    wb = openpyxl.load_workbook('Marcadores Mundial 2026.xlsx', data_only=True)
    res_sheet = wb['Resultados']
    
    res_matches = []
    for r in range(2, res_sheet.max_row + 1):
        g = res_sheet.cell(r, 1).value
        local = res_sheet.cell(r, 2).value
        visitante = res_sheet.cell(r, 6).value
        if g and local and visitante:
            res_matches.append((g, local.strip(), visitante.strip(), r))
            
    print(f"Found {len(res_matches)} matches in Resultados sheet.")
    
    participants = ['Deivid', 'Cesar', 'Ferney', 'Edisabet', 'Jhair', 'Duvan']
    for p in participants:
        p_sheet = wb[p]
        p_matches = []
        for r in range(2, p_sheet.max_row + 1):
            g = p_sheet.cell(r, 1).value
            local = p_sheet.cell(r, 2).value
            visitante = p_sheet.cell(r, 4).value
            pred = p_sheet.cell(r, 5).value
            if g and local and visitante:
                p_matches.append((g, local.strip(), visitante.strip(), pred, r))
        
        print(f"Participant {p}: {len(p_matches)} matches in sheet.")
        
        # Compare matches
        mismatches = 0
        for idx, (rg, rl, rv, rr) in enumerate(res_matches):
            if idx >= len(p_matches):
                mismatches += 1
                continue
            pg, pl, pv, pp, pr = p_matches[idx]
            if rg != pg or rl != pl or rv != pv:
                print(f"  Mismatch at index {idx}: Resultados=({rg}, {rl}, {rv}) vs {p}=({pg}, {pl}, {pv})")
                mismatches += 1
        
        if mismatches == 0:
            print(f"  Matches align perfectly for {p}!")
        else:
            print(f"  Total mismatches for {p}: {mismatches}")

if __name__ == '__main__':
    main()
