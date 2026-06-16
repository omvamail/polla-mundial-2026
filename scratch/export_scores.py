import json
import openpyxl

def main():
    wb = openpyxl.load_workbook('Marcadores Mundial 2026.xlsx', data_only=True)
    res_sheet = wb['Resultados']
    
    scores = {}
    for r in range(2, res_sheet.max_row + 1):
        g = res_sheet.cell(r, 1).value
        local = res_sheet.cell(r, 2).value
        goles_l = res_sheet.cell(r, 3).value
        goles_v = res_sheet.cell(r, 5).value
        visitante = res_sheet.cell(r, 6).value
        
        if g and local and visitante:
            if goles_l is not None and goles_v is not None:
                # Store by row number as a string key in JSON
                scores[str(r)] = {
                    'goles_local': int(goles_l),
                    'goles_visitante': int(goles_v)
                }
                
    with open('scores.json', 'w', encoding='utf-8') as f:
        json.dump(scores, f, indent=4, ensure_ascii=False)
        
    print(f"Exported {len(scores)} played match scores to scores.json!")

if __name__ == '__main__':
    main()
