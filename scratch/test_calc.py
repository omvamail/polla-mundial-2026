import openpyxl

def calculate_scores():
    wb = openpyxl.load_workbook('Marcadores Mundial 2026.xlsx', data_only=True)
    res_sheet = wb['Resultados']
    
    # 1. Read matches and results
    matches = {}
    for r in range(2, res_sheet.max_row + 1):
        g = res_sheet.cell(r, 1).value
        local = res_sheet.cell(r, 2).value
        goles_l = res_sheet.cell(r, 3).value
        goles_v = res_sheet.cell(r, 5).value
        visitante = res_sheet.cell(r, 6).value
        outcome = res_sheet.cell(r, 7).value # Outcome can be 'Local (L)', 'Visitante (V)', 'Empate (E)' or None
        
        if g and local and visitante:
            # Clean spaces
            local = local.strip()
            visitante = visitante.strip()
            matches[r] = {
                'group': g,
                'local': local,
                'goles_local': goles_l,
                'goles_visitante': goles_v,
                'visitante': visitante,
                'outcome': outcome.strip() if outcome else None
            }
            
    # 2. Check each participant
    participants = ['Deivid', 'Cesar', 'Ferney', 'Edisabet', 'Jhair', 'Duvan']
    calculated_scores = {}
    
    for p in participants:
        p_sheet = wb[p]
        pts = 0
        correct_matches = 0
        details = []
        
        # Read the Excel total points (cell J1)
        excel_total = p_sheet.cell(1, 10).value
        
        for r, m in matches.items():
            # In participant sheets, Column 5 is 'Predicción'
            pred = p_sheet.cell(r, 5).value
            if pred:
                pred = pred.strip()
            
            # Check correctness
            is_correct = False
            points_awarded = 0
            if m['outcome'] and pred:
                # Excel checks: IF(IFERROR(FIND(Resultados!G2,Deivid!E2),0)>0,"🟢","🔴")
                # So if outcome is in prediction string, it's correct
                if m['outcome'] in pred:
                    is_correct = True
                    points_awarded = 1
                    pts += 1
                    correct_matches += 1
            
            details.append({
                'row': r,
                'match': f"{m['local']} vs {m['visitante']}",
                'prediction': pred,
                'outcome': m['outcome'],
                'correct': is_correct,
                'points': points_awarded
            })
            
        calculated_scores[p] = {
            'calculated_points': pts,
            'excel_total': excel_total,
            'correct_matches': correct_matches
        }
        
    return calculated_scores

if __name__ == '__main__':
    scores = calculate_scores()
    print("Scoring Verification:")
    for p, s in scores.items():
        print(f"Participant: {p}")
        print(f"  Calculated: {s['calculated_points']}")
        print(f"  Excel Total: {s['excel_total']}")
        print(f"  Correct guesses: {s['correct_matches']}")
        match = s['calculated_points'] == s['excel_total']
        print(f"  Matches? {'✅ YES' if match else '❌ NO'}")
