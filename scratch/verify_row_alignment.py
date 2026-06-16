import openpyxl

wb = openpyxl.load_workbook('Marcadores Mundial 2026.xlsx', data_only=True)
res_sheet = wb['Resultados']
participants = ['Deivid', 'Cesar', 'Ferney', 'Edisabet', 'Jhair', 'Duvan']

# Let's check all row indices from 2 to 90
mismatches_found = False
for r in range(2, res_sheet.max_row + 1):
    res_g = res_sheet.cell(r, 1).value
    res_l = res_sheet.cell(r, 2).value
    res_v = res_sheet.cell(r, 6).value
    
    # Check if this is a match row
    if res_g or res_l or res_v:
        for p in participants:
            p_sheet = wb[p]
            p_g = p_sheet.cell(r, 1).value
            p_l = p_sheet.cell(r, 2).value
            p_v = p_sheet.cell(r, 4).value
            
            # Normalize names to check alignment
            def norm(name):
                if not name: return ""
                n = str(name).strip().lower()
                if n == "holanda": return "países bajos"
                return n
            
            if norm(res_g) != norm(p_g) or norm(res_l) != norm(p_l) or norm(res_v) != norm(p_v):
                print(f"Row {r} mismatch for {p}: Resultados=({res_g}, {res_l}, {res_v}) vs {p}=({p_g}, {p_l}, {p_v})")
                mismatches_found = True

if not mismatches_found:
    print("Perfect row alignment across all sheets and all participants (with 'Holanda' normalized to 'Países Bajos')!")
else:
    print("Some row mismatches were found.")
