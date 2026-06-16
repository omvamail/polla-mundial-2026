import openpyxl

wb = openpyxl.load_workbook('Marcadores Mundial 2026.xlsx', data_only=False)
for name in ['Deivid', 'Cesar', 'Ferney', 'Edisabet', 'Jhair', 'Duvan']:
    sheet = wb[name]
    print(f"Sheet {name}: cell J1 = {sheet.cell(1, 10).value}, cell I1 = {sheet.cell(1, 9).value}")
