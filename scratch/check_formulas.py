import openpyxl

wb = openpyxl.load_workbook('Marcadores Mundial 2026.xlsx', data_only=False)
sheet = wb['Resultados']

print("Cell formulas / values in Resultados sheet:")
# Check rows 1-3
for r in range(1, 4):
    row_vals = [sheet.cell(r, c).value for c in range(1, 12)]
    print(f"Row {r}: {row_vals}")

# Also check Deivid sheet row 1-3
print("\nCell formulas / values in Deivid sheet:")
d_sheet = wb['Deivid']
for r in range(1, 4):
    row_vals = [d_sheet.cell(r, c).value for c in range(1, 10)]
    print(f"Row {r}: {row_vals}")
